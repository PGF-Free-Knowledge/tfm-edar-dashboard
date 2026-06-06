from openpyxl import load_workbook

archivo = r"data/raw/TFM_EDAR_3.xlsx"

wb = load_workbook(archivo, data_only=False)
ws = wb["Cerceda Balance"]

celdas = [
    "AB157",
    "AB165",
    "S176"
]

print("\nSIGUIENDO AGUAS ARRIBA\n")

for c in celdas:
    print(f"{c:6} -> {ws[c].value}")