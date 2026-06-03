from openpyxl import load_workbook

archivo = "data/raw/TFM_EDAR_2.xlsx"

wb = load_workbook(archivo, data_only=True)

print("\nBUSCANDO VALORES NEGATIVOS\n")

for hoja in wb.worksheets:
    encontrados = 0

    for fila in hoja.iter_rows():
        for celda in fila:

            if isinstance(celda.value, (int, float)):
                if celda.value < 0:
                    encontrados += 1
                    print(
                        f"Hoja: {hoja.title:20} "
                        f"Celda: {celda.coordinate:8} "
                        f"Valor: {celda.value}"
                    )

    if encontrados > 0:
        print(f"\nTotal negativos en {hoja.title}: {encontrados}\n")