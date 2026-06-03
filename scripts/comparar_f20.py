from openpyxl import load_workbook

archivo_a = "data/raw/TFM_EDAR_2.xlsx"
archivo_b = "data/raw/TFM_EDAR_2_F20_act.xlsx"

wb_a = load_workbook(archivo_a, data_only=True)
wb_b = load_workbook(archivo_b, data_only=True)

hoja = "Cerceda Balance"

ws_a = wb_a[hoja]
ws_b = wb_b[hoja]

print("\nCELDAS QUE CAMBIARON\n")

cambios = 0

for fila in range(1, ws_a.max_row + 1):
    for col in range(1, ws_a.max_column + 1):

        v1 = ws_a.cell(fila, col).value
        v2 = ws_b.cell(fila, col).value

        if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):

            if abs(v1 - v2) > 1e-9:

                cambios += 1

                print(
                    f"{ws_a.cell(fila,col).coordinate:8} | "
                    f"{v1:12.4f} -> {v2:12.4f}"
                )

print(f"\nTOTAL CAMBIOS: {cambios}")