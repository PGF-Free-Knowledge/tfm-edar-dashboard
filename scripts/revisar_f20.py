from openpyxl import load_workbook

archivo = "data/raw/TFM_EDAR_2.xlsx"

wb = load_workbook(archivo, data_only=True)
ws = wb["Cerceda Balance"]

celdas = [
    "S176",
    "AB165",
    "AB157",
    "R368",
    "R369"
]
print("\nREVISIÓN DE CELDAS\n")

for celda in celdas:
    print(
        f"{celda:6} | "
        f"valor={ws[celda].value}"
    )